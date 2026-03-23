"""Verify album records in the Supabase database using the Claude Agent SDK.

Fetches all album records via the Supabase Python client, then sends
the data to Claude Code for verification of each record's consistency
(title, artist, release year, label, summaries, etc.). Writes findings
to an xlsx file.

Albums are split into batches and verified in parallel for speed.
"""

import asyncio
import json
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from supabase import Client, create_client

from claude_agent_sdk import (
    AssistantMessage,
    ClaudeAgentOptions,
    ResultMessage,
    TextBlock,
    ToolUseBlock,
    query,
)

SPINNER_FRAMES = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
BATCH_SIZE = 50
MAX_CONCURRENT = 4

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
BACKEND_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BACKEND_ROOT / "output"

COLUMNS_TO_FETCH = (
    "album_id, title, artist, release_year, label_name, cover_artists, "
    "artist_summary, album_summary, streaming_link_spotify, streaming_link_apple, "
    "cover_image_url, calendar_order, apple_link_is_substitute, "
    "spotify_link_is_substitute, image_filename"
)

# ANSI colors
_GREEN = "\033[1;32m"
_YELLOW = "\033[0;33m"
_RED = "\033[0;31m"
_DIM = "\033[2m"
_RESET = "\033[0m"


class AlbumFinding(BaseModel):
    album_id: str
    title: str
    artist: str
    is_legitimate: bool
    release_year_correct: bool
    label_correct: bool
    artist_summary_coherent: bool
    album_summary_coherent: bool
    cover_artists_correct: bool
    apple_link_coherent: bool
    issues: str
    confidence: str
    image_filename: str | None


class VerificationResults(BaseModel):
    findings: list[AlbumFinding]


VERIFICATION_PROMPT_TEMPLATE = """\
You are a jazz music expert verifying the accuracy of album records in a database.

Below are ALL the album records from the database, provided as JSON. You do NOT need
to query any database or use any tools — the data is right here.

<album_records>
{records_json}
</album_records>

For EACH album record, verify the following:

1. **is_legitimate**: Is this title + artist combination a real, legitimate jazz album?
   Mark false if the album doesn't exist or is clearly wrong.

2. **release_year_correct**: Does the release_year in the database match the actual
   release year of this album? If release_year is NULL, mark as true but note it in issues.

3. **label_correct**: Does the label_name match the actual record label for this album?
   If label_name is NULL, mark as true but note it in issues.

4. **artist_summary_coherent**: Is the artist_summary actually about the correct artist
   and factually reasonable? If NULL, mark as true but note it.

5. **album_summary_coherent**: Is the album_summary actually about the correct album
   and factually reasonable? If NULL, mark as true but note it.

6. **cover_artists_correct**: Are the cover_artists (sidemen/personnel) coherent with
   this album? If NULL, mark as true but note it.

7. **apple_link_coherent**: Check the streaming_link_apple URL text (do NOT visit the URL).
   Apple Music URLs typically contain the artist name, album title, and sometimes a year
   in their path (e.g., "https://music.apple.com/us/album/blue-train/724467923").
   Verify that the text in the URL is coherent with the album's title and artist.
   If the URL contains a different artist or album name, mark as false.
   If streaming_link_apple is NULL, mark as true but note it in issues.
   Also consider apple_link_is_substitute: if true, the link may point to a different
   album by the same artist (which is acceptable but should be noted).

8. **issues**: A free-text description of ALL issues found. Include:
   - Any factual inaccuracies
   - NULL columns that should have values
   - Mismatched data (e.g., wrong year, wrong label)
   - Streaming links that look malformed (not full URLs)
   - Any other inconsistencies
   If no issues, use "No issues found."

9. **confidence**: Your confidence in the verification: "high", "medium", or "low".
   Use "low" if you're unsure about the album's existence or details.

10. **image_filename**: Copy the image_filename value directly from the album record
    (may be NULL).

Return the results as structured JSON matching the required schema.
You MUST return one finding per album record. Do not skip any records.
"""


@dataclass
class BatchState:
    batch_num: int
    total_batches: int
    start_idx: int   # index into the full albums list
    count: int       # number of albums in this batch
    status: str = "queued"  # queued | running | done | error
    turn: int = 0
    message: str = ""
    last_activity: float = field(default_factory=time.monotonic)
    start_time: float = 0.0
    elapsed: float = 0.0
    spinner_idx: int = 0
    cost: float = 0.0
    findings_count: int = 0
    error: str = ""

    @property
    def label(self) -> str:
        end = self.start_idx + self.count
        return f"albums {self.start_idx + 1}–{end}"


def _render_batch_line(state: BatchState, term_width: int = 100) -> str:
    prefix = f"  [{state.batch_num:2d}/{state.total_batches}]"

    if state.status == "queued":
        return f"{prefix} {_DIM}{state.label} — queued{_RESET}"

    if state.status == "running":
        frame = SPINNER_FRAMES[state.spinner_idx % len(SPINNER_FRAMES)]
        elapsed = time.monotonic() - state.start_time
        mins, secs = divmod(int(elapsed), 60)
        ts = f"{mins:02d}:{secs:02d}"
        idle = int(time.monotonic() - state.last_activity)
        idle_str = f" {_DIM}({idle}s ago){_RESET}" if idle >= 10 else ""
        msg = state.message
        max_msg = 45
        if len(msg) > max_msg:
            msg = msg[:max_msg] + "…"
        return f"{prefix} {_YELLOW}{frame}{_RESET} [{ts}] Turn {state.turn}: {msg}{idle_str}"

    if state.status == "done":
        mins, secs = divmod(int(state.elapsed), 60)
        cost_str = f"  ${state.cost:.4f}" if state.cost else ""
        return (
            f"{prefix} {_GREEN}✓{_RESET} Done in {mins}m {secs}s"
            f" — {state.findings_count} findings{cost_str}"
        )

    if state.status == "error":
        err = state.error[:55] if len(state.error) > 55 else state.error
        return f"{prefix} {_RED}✗ Error: {err}{_RESET}"

    return f"{prefix} {state.status}"


async def _display_loop(
    states: list[BatchState],
    total_albums: int,
    stop_event: asyncio.Event,
) -> None:
    """Redraws all batch status lines every 0.2 seconds until stop_event is set."""
    n_batch_lines = len(states)
    n_total_lines = n_batch_lines + 2  # batches + blank + progress

    # Reserve space
    for _ in range(n_total_lines):
        print()

    while True:
        sys.stdout.write(f"\033[{n_total_lines}A")

        for state in states:
            sys.stdout.write(f"\033[K{_render_batch_line(state)}\n")

        done = sum(1 for s in states if s.status == "done")
        errors = sum(1 for s in states if s.status == "error")
        albums_done = sum(s.count for s in states if s.status in ("done", "error"))
        running = sum(1 for s in states if s.status == "running")
        err_str = f"  {_RED}{errors} error(s){_RESET}" if errors else ""

        sys.stdout.write("\033[K\n")
        sys.stdout.write(
            f"\033[K  Progress: {done}/{len(states)} batches done"
            f" | {albums_done}/{total_albums} albums"
            f" | {running} running{err_str}\n"
        )
        sys.stdout.flush()

        if stop_event.is_set():
            break

        await asyncio.sleep(0.2)


async def _verify_batch(
    albums: list[dict],
    state: BatchState,
    semaphore: asyncio.Semaphore,
) -> VerificationResults:
    """Verify one batch of albums, updating state for live display."""
    async with semaphore:
        state.status = "running"
        state.start_time = time.monotonic()
        state.last_activity = time.monotonic()
        state.message = "Starting…"

        records_json = json.dumps(albums, indent=2, ensure_ascii=False)
        prompt = VERIFICATION_PROMPT_TEMPLATE.format(records_json=records_json)

        options = ClaudeAgentOptions(
            cwd=str(PROJECT_ROOT),
            permission_mode="bypassPermissions",
            disallowed_tools=["Bash", "Edit", "Write"],
            output_format={
                "type": "json_schema",
                "schema": VerificationResults.model_json_schema(),
            },
            max_turns=50,
        )

        result_text = None
        structured_output = None

        async for message in query(prompt=prompt, options=options):
            state.spinner_idx += 1
            state.last_activity = time.monotonic()

            if isinstance(message, AssistantMessage):
                state.turn += 1
                for block in message.content:
                    if isinstance(block, ToolUseBlock):
                        state.message = f"Calling {block.name}…"
                    elif isinstance(block, TextBlock):
                        text = block.text.replace("\n", " ").strip()
                        if text:
                            state.message = text
            elif isinstance(message, ResultMessage):
                state.elapsed = time.monotonic() - state.start_time
                state.cost = message.total_cost_usd or 0.0
                if message.is_error:
                    state.status = "error"
                    state.error = message.subtype or "unknown error"
                    return VerificationResults(findings=[])
                result_text = message.result
                structured_output = message.structured_output

        if structured_output:
            results = VerificationResults.model_validate(structured_output)
        elif result_text:
            try:
                data = json.loads(result_text)
                results = VerificationResults.model_validate(data)
            except Exception as exc:
                state.status = "error"
                state.error = str(exc)
                return VerificationResults(findings=[])
        else:
            state.status = "error"
            state.error = "No results returned"
            return VerificationResults(findings=[])

        state.findings_count = len(results.findings)
        state.status = "done"
        return results


async def verify_albums(albums: list[dict]) -> VerificationResults:
    """Split albums into batches and verify them in parallel."""
    batches = [albums[i : i + BATCH_SIZE] for i in range(0, len(albums), BATCH_SIZE)]
    total_batches = len(batches)

    states = [
        BatchState(
            batch_num=i + 1,
            total_batches=total_batches,
            start_idx=i * BATCH_SIZE,
            count=len(batch),
        )
        for i, batch in enumerate(batches)
    ]

    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    stop_event = asyncio.Event()

    print()
    print(
        f"  Verifying {len(albums)} albums in {total_batches} batches"
        f" (up to {MAX_CONCURRENT} concurrent)…"
    )

    display_task = asyncio.create_task(_display_loop(states, len(albums), stop_event))
    batch_tasks = [
        asyncio.create_task(_verify_batch(batch, states[i], semaphore))
        for i, batch in enumerate(batches)
    ]

    batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)

    stop_event.set()
    await display_task

    all_findings: list[AlbumFinding] = []
    for i, result in enumerate(batch_results):
        if isinstance(result, Exception):
            print(f"\n  Batch {i + 1} raised an exception: {result}")
        elif isinstance(result, VerificationResults):
            all_findings.extend(result.findings)

    print()
    total_cost = sum(s.cost for s in states)
    if total_cost:
        print(f"  Total cost: ${total_cost:.4f}")

    if not all_findings:
        print("  Error: No findings returned from any batch.")
        sys.exit(1)

    return VerificationResults(findings=all_findings)


def get_supabase_client() -> Client:
    """Initialize Supabase client from environment variables."""
    load_dotenv(os.path.join(BACKEND_ROOT, ".env.local"))
    if os.environ.get("PRODUCTION") == "True":
        load_dotenv(os.path.join(BACKEND_ROOT, ".env.production"), override=True)
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("Error: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set.")
        sys.exit(1)
    return create_client(url, key)


def fetch_albums(client: Client) -> list[dict]:
    """Fetch all album records from Supabase."""
    response = client.table("albums").select(COLUMNS_TO_FETCH).execute()
    return response.data


def write_xlsx(results: VerificationResults, output_path: Path) -> None:
    """Write verification findings to an xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Album Verification"

    headers = [
        "Album ID",
        "Title",
        "Artist",
        "Legitimate?",
        "Release Year OK?",
        "Label OK?",
        "Artist Summary OK?",
        "Album Summary OK?",
        "Cover Artists OK?",
        "Apple Link OK?",
        "Issues",
        "Confidence",
        "Image Filename",
    ]

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    bool_columns = [4, 5, 6, 7, 8, 9, 10]

    for row_idx, finding in enumerate(results.findings, 2):
        row_data = [
            finding.album_id,
            finding.title,
            finding.artist,
            finding.is_legitimate,
            finding.release_year_correct,
            finding.label_correct,
            finding.artist_summary_coherent,
            finding.album_summary_coherent,
            finding.cover_artists_correct,
            finding.apple_link_coherent,
            finding.issues,
            finding.confidence,
            finding.image_filename,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in bool_columns:
                cell.value = "YES" if value else "NO"
                cell.fill = green_fill if value else red_fill
            if col_idx == 11:
                cell.alignment = wrap_alignment

    column_widths = {
        "A": 38,
        "B": 35,
        "C": 25,
        "D": 13,
        "E": 17,
        "F": 12,
        "G": 19,
        "H": 19,
        "I": 17,
        "J": 16,
        "K": 60,
        "L": 13,
        "M": 35,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(results.findings) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Wrote {len(results.findings)} findings to {output_path}")


def print_summary(results: VerificationResults) -> None:
    """Print a summary of the verification results."""
    total = len(results.findings)
    issues_count = sum(1 for f in results.findings if f.issues != "No issues found.")
    illegitimate = sum(1 for f in results.findings if not f.is_legitimate)
    low_confidence = sum(1 for f in results.findings if f.confidence == "low")

    print()
    print("=" * 60)
    print("  VERIFICATION SUMMARY")
    print("=" * 60)
    print(f"  Total albums verified: {total}")
    print(f"  Albums with issues:    {issues_count}")
    print(f"  Illegitimate albums:   {illegitimate}")
    print(f"  Low confidence:        {low_confidence}")
    print("=" * 60)


async def main() -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_path = OUTPUT_DIR / f"album-verification-{timestamp}.xlsx"

    print("=" * 60)
    print("  Album Record Verification")
    print("  Supabase -> Claude Code -> XLSX")
    print("=" * 60)

    print()
    print("  Connecting to Supabase...")
    client = get_supabase_client()
    albums = fetch_albums(client)
    print(f"  Fetched {len(albums)} album records.")

    if not albums:
        print("  No albums found. Nothing to verify.")
        sys.exit(0)

    # Remove ANTHROPIC_API_KEY so the Agent SDK uses Claude Code's
    # subscription auth instead of a potentially invalid API key.
    os.environ.pop("ANTHROPIC_API_KEY", None)

    results = await verify_albums(albums)

    write_xlsx(results, output_path)
    print_summary(results)


if __name__ == "__main__":
    asyncio.run(main())
